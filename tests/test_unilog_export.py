import io
import openpyxl
import pytest

from app.pipeline import export_unilog_excel_bytes, synthesize_5tier_descriptions


def test_synthesize_5tier_descriptions():
    part = "ACS580-01-018A-4"
    mfg = "ABB"
    cat = "Industrial Automation > Drives > Variable Frequency Drives (VFD)"
    field_map = {
        "voltage_rating": {"value": "400", "unit": "V AC"},
        "power_rating": {"value": "7.5", "unit": "kW"},
        "current_rating": {"value": "17.7", "unit": "A"},
        "ip_rating": {"value": "IP21", "unit": ""},
    }

    descs = synthesize_5tier_descriptions(part, mfg, cat, field_map)

    # 1. Mobile description <= 80 chars
    assert len(descs["mobile"]) <= 80
    assert "ABB" in descs["mobile"]
    assert "ACS580" in descs["mobile"]

    # 2. In-search description <= 150 chars
    assert len(descs["search"]) <= 150

    # 3. Short description <= 250 chars
    assert len(descs["short"]) <= 250

    # 4. Long description has bulleted structure
    assert "• Manufacturer:" in descs["long"]
    assert "• Part Number:" in descs["long"]

    # 5. Marketing description contains verbatim OEM wording
    assert "Official ABB" in descs["marketing"]
    assert "corroborated" in descs["marketing"]


def test_export_unilog_excel_bytes_columns():
    sample_products = [
        {
            "part_number": "ATV320U07N4B",
            "manufacturer": "Schneider Electric",
            "category": "Industrial Automation > Drives > Variable Frequency Drives (VFD)",
            "Mobile_Description": "Schneider ATV320 0.75kW 400V IP20 VFD",
            "fields": [
                {"attribute_key": "voltage_rating", "value": "400", "unit": "V AC"},
                {"attribute_key": "power_rating", "value": "0.75", "unit": "kW"},
                {"attribute_key": "current_rating", "value": "2.3", "unit": "A"},
                {"attribute_key": "ip_rating", "value": "IP20", "unit": ""},
                {"attribute_key": "operating_temp", "value": "-10 to 50", "unit": "°C"},
            ],
            "Datasheet_PDF_URL": "https://www.se.com/datasheets/ATV320.pdf",
            "Provenance_Source_URL": "https://www.se.com/products/ATV320",
        },
        {
            "part_number": "6SL3210-1KE11-8UB1",
            "manufacturer": "Siemens",
            "category": "Industrial Automation > Drives > Variable Frequency Drives (VFD)",
            "fields": [
                {"attribute_key": "voltage_rating", "value": "400", "unit": "V AC"},
                {"attribute_key": "power_rating", "value": "0.55", "unit": "kW"},
                {"attribute_key": "current_rating", "value": "1.7", "unit": "A"},
                {"attribute_key": "ip_rating", "value": "IP20", "unit": ""},
            ],
        }
    ]

    xlsx_data = export_unilog_excel_bytes(sample_products)
    assert isinstance(xlsx_data, bytes)
    assert len(xlsx_data) > 0

    # Read back the workbook
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_data))
    ws = wb.active
    assert ws.title == "Unilog Enriched Catalog"

    headers = [cell.value for cell in ws[1]]
    
    # Check mandatory Unilog columns
    assert "Manufacturer_Part_Number" in headers
    assert "Manufacturer_Name" in headers
    assert "Taxonomy_Leaf_Category" in headers
    assert "Mobile_Description" in headers
    assert "In_Search_Description" in headers
    assert "Short_Description" in headers
    assert "Long_Description" in headers
    assert "Marketing_Description" in headers
    assert "Primary_Image_URL" in headers
    assert "Datasheet_PDF_URL" in headers
    assert "Provenance_Source_URL" in headers

    # Check dynamic Value and UOM split columns
    assert "Attr_VoltageRating_Value" in headers
    assert "Attr_VoltageRating_UOM" in headers
    assert "Attr_PowerRating_Value" in headers
    assert "Attr_PowerRating_UOM" in headers

    # Verify Row Data
    row2 = {headers[i]: ws.cell(row=2, column=i+1).value for i in range(len(headers))}
    assert row2["Manufacturer_Part_Number"] == "ATV320U07N4B"
    assert row2["Manufacturer_Name"] == "Schneider Electric"
    assert row2["Taxonomy_Leaf_Category"] == "Industrial Automation > Drives > Variable Frequency Drives (VFD)"
    assert row2["Attr_VoltageRating_Value"] == "400"
    assert row2["Attr_VoltageRating_UOM"] == "V AC"
    assert row2["Attr_PowerRating_Value"] == "0.75"
    assert row2["Attr_PowerRating_UOM"] == "kW"

    row3 = {headers[i]: ws.cell(row=3, column=i+1).value for i in range(len(headers))}
    assert row3["Manufacturer_Part_Number"] == "6SL3210-1KE11-8UB1"
    assert row3["Manufacturer_Name"] == "Siemens"
